import openpyxl
from openpyxl import Workbook
import os as os
from xlsconverter import cvt_xls_to_xlsx
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment
firstcolumn = Font(bold = True , size =20, color = colors.BLACK)
midcolumn = Font(bold = True , size =24, color = colors.BLACK)
while True:
    #taking file path src and dst from user
    srcfname = input('enter the sorcue file path without double quotattions ')
    desfname = srcfname + 'x'
    sitename = input('enter the site name')
    blance_amount = input('enter the blance amount')
    #converting from xls to xlsx
    cvt_xls_to_xlsx(srcfname, desfname)

    ws = openpyxl.load_workbook(desfname) 
    sheet= ws.active
    max_column = sheet.max_column
    max_row  = sheet.max_row
    print('Edditing the file ...')
    for row in sheet.iter_rows(min_row=1,min_col = 1 ,max_col = 1):
            for cell in row:
                cell.value = "T R X"
                cell.font = Font(size = 22 , bold = True)
                sheet.column_dimensions['A'].width = 21
                cell.alignment= Alignment(horizontal = 'center',vertical= 'center')

    for col in ['H', 'I', 'F','G']:
        sheet.column_dimensions[col].hidden= True

    sheet.insert_cols(2)
    sheet.insert_cols(4)

    for i in range(2,max_row+1):
        # iterate over all columns
        sheet.cell(row = 1,column = 4).value = 'Balance'
        cell_val = sheet.cell(row = i,column = 5).value
        cell_sum  = cell_val/1024
        cell_sum= cell_sum/1024
        sheet.cell(row = i,column = 4).value = '{} MB'.format(int(cell_sum))
        sheet.cell(row = i,column = 4).font = Font(size=22 ,bold= True ,vertAlign = 'subscript')
        sheet.cell(row = 1,column = 4).font = Font(size=22 ,bold= True ,vertAlign = 'subscript')
        sheet.cell(row = i,column = 4).alignment = Alignment(horizontal = 'center',vertical= 'center')
        sheet.cell(row = 1,column = 4).alignment = Alignment(horizontal = 'center',vertical= 'center')
        sheet.column_dimensions['D'].width = 22





    for k in range(2,max_row):
        sheet.cell(row = k,column = 2).value = '=text(C%d," 0 0 0 0 0 0 0 0 ")' % (k)
        sheet.cell(row = 1,column = 2).value = 'PIN Code'
        sheet.cell(row = k,column = 2).font = Font(size=24 ,bold= True )
        sheet.cell(row = 1,column = 2).font = Font(size=25 ,bold= True ,vertAlign = 'subscript')
        sheet.cell(row = 1,column = 2).alignment = Alignment(horizontal = 'center',vertical= 'center')
        sheet.cell(row = k,column = 2).alignment = Alignment(horizontal = 'center',vertical= 'center')
        sheet.row_dimensions[k].height = 150
        sheet.column_dimensions['B'].width = 30

    for col in ['C','E']:
        sheet.column_dimensions[col].hidden= True
    current_dir = os.getcwd()
    print("Done, you can go to ")
    print(current_dir)
    print(" to see your new  file") 
    new_file = '{}-{}.xlsx'.format(sitename,blance_amount)

    ws.save(new_file)
